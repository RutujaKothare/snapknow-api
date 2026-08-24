"""
AI Image Caption & Info Assistant
----------------------------------
Upload an image (animal, flower, food, vegetable, or anything else) and this app will:
  1. Caption the image and identify it using Claude's vision (an LLM API).
  2. Give detailed information about it (habitat/diet for animals, nutrition for food, etc).
  3. Read the result aloud (Text-to-Speech).
  4. Let you ask follow-up questions by voice (Speech-to-Text).
  5. Let you download a PDF report of the analysis.

Run with:  streamlit run app.py
"""

import streamlit as st
import base64
from PIL import Image
import io
from gtts import gTTS
import speech_recognition as sr
from fpdf import FPDF
import os
import json
import csv
import re
import hashlib
import logging
import urllib.request
import urllib.parse
import urllib.error
import tempfile
from datetime import datetime
from dotenv import load_dotenv

from core import (
    get_media_type,
    guess_media_type_from_bytes,
    resize_image_bytes,
    ANALYSIS_PROMPT,
    analyze_image,
    translate_result_text,
    text_completion,
    parse_response,
)
import rag_engine

# ---------------------------------------------------------------------------
# .env support — lets API keys be set once as environment variables instead
# of retyped into the UI every run. Sidebar fields still work as an override
# (and as the only option if you'd rather not use a .env file at all).
# ---------------------------------------------------------------------------
load_dotenv()

# ---------------------------------------------------------------------------
# Logging — every previously-silent "except: pass" now logs what happened,
# so failures are visible in the terminal instead of disappearing quietly.
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("snapknow.app")

st.set_page_config(page_title="SnapKnow — AI Image Identifier", page_icon="📸", layout="wide")

# ---------------------------------------------------------------------------
# Visual theme — playful photo-scrapbook look. Results render as tilted
# polaroid cards with a sticker-style category tag and washi-tape corners;
# history becomes a small gallery of pinned thumbnails. Fits an app whose
# whole job is "you snapped a photo, here's what it is."
# ---------------------------------------------------------------------------
CATEGORY_STYLES = {
    "animal":    {"emoji": "🐾", "bg": "#ECFDF3", "text": "#027A48"},
    "flower":    {"emoji": "🌸", "bg": "#FDF2FA", "text": "#C11574"},
    "food":      {"emoji": "🍽️", "bg": "#FFF6ED", "text": "#C4320A"},
    "vegetable": {"emoji": "🥕", "bg": "#F7FEE7", "text": "#4D7C0F"},
    "fruit":     {"emoji": "🍎", "bg": "#FEF3F2", "text": "#B42318"},
    "object":    {"emoji": "📦", "bg": "#EFF8FF", "text": "#175CD3"},
    "place":     {"emoji": "🗺️", "bg": "#F4F3FF", "text": "#5925DC"},
    "person":    {"emoji": "🧑", "bg": "#F0FDFA", "text": "#0E7490"},
}
DEFAULT_STYLE = {"emoji": "✦", "bg": "#F8FAFC", "text": "#475467"}

# ---------------------------------------------------------------------------
# Multi-language support — one code drives both the LLM prompt language and
# the gTTS voice used for "Read Aloud".
# ---------------------------------------------------------------------------
LANGUAGE_OPTIONS = {
    "English": "en",
    "Hindi (हिन्दी)": "hi",
    "Marathi (मराठी)": "mr",
    "Spanish (Español)": "es",
    "French (Français)": "fr",
}

# ---------------------------------------------------------------------------
# Local history persistence — saves to a JSON file next to app.py so your
# analyzed images survive a browser refresh or restarting the app. Defined
# early since session-state init (below) needs it right away.
# ---------------------------------------------------------------------------
HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "snapknow_history.json")


def load_history_from_disk() -> list:
    if not os.path.exists(HISTORY_FILE):
        return []
    try:
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        entries = []
        for entry in raw:
            entries.append({
                "result": entry["result"],
                "image_bytes": base64.b64decode(entry["image_b64"]),
                # Older history entries were saved before timestamp/language were
                # tracked — default them clearly rather than leaving a blank/None
                # that would show up as a null cell in any export.
                "timestamp": entry.get("timestamp") or "Unknown",
                "language": entry.get("language") or "English",
            })
        return entries
    except Exception:
        logger.error(f"Could not load history from {HISTORY_FILE}; starting with empty history", exc_info=True)
        return []  # corrupted or unreadable file — start fresh rather than crashing


def save_history_to_disk(history: list):
    try:
        serializable = [
            {
                "result": entry["result"],
                "image_b64": base64.b64encode(entry["image_bytes"]).decode("ascii"),
                "timestamp": entry.get("timestamp") or "Unknown",
                "language": entry.get("language") or "English",
            }
            for entry in history
        ]
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(serializable, f)
    except Exception:
        logger.error(f"Could not save history to {HISTORY_FILE} — changes will be lost on restart", exc_info=True)


# ---------------------------------------------------------------------------
# CSV export — writes one row per analyzed image, plus saves each photo as an
# actual file (CSV itself can't hold binary image data cleanly). The CSV
# references each image by filename so you can open both side by side.
# ---------------------------------------------------------------------------
IMAGES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "snapknow_images")

CSV_COLUMNS = [
    "image_filename", "caption", "category", "name", "confidence",
    "description", "details", "fun_fact",
    "wikipedia_summary", "wikipedia_url",
    "nutrition_calories", "nutrition_protein", "nutrition_carbs", "nutrition_fat",
    "gbif_scientific_name", "gbif_kingdom", "gbif_family", "gbif_status",
    "recipe_ingredients", "recipe_instructions",
    "weather_description", "weather_temp_c",
]

# Human-readable column headers, in the same order as CSV_COLUMNS, for the Excel export.
XLSX_HEADERS = [
    "Image", "Caption", "Category", "Name", "Confidence",
    "Description", "Details", "Fun Fact",
    "Wikipedia Summary", "Wikipedia URL",
    "Calories", "Protein", "Carbs", "Fat",
    "Scientific Name", "Kingdom", "Family", "Conservation Status",
    "Recipe Ingredients", "Recipe Instructions",
    "Weather", "Temp (C)",
]


def _save_history_image(image_bytes: bytes) -> str:
    """Save one history entry's image to IMAGES_DIR (skipping ones already saved,
    identified by content hash) and return its filename."""
    os.makedirs(IMAGES_DIR, exist_ok=True)
    img_hash = hashlib.md5(image_bytes).hexdigest()[:12]
    ext = guess_media_type_from_bytes(image_bytes).split("/")[-1]
    filename = f"{img_hash}.{ext}"
    filepath = os.path.join(IMAGES_DIR, filename)
    if not os.path.exists(filepath):
        try:
            with open(filepath, "wb") as f:
                f.write(image_bytes)
        except Exception:
            logger.error(f"Could not save image file {filepath} for export", exc_info=True)
            return "(could not save)"
    return filename


def _clean_ingredient_line(text: str) -> str:
    """Collapse stray whitespace inside a single ingredient line (MealDB
    sometimes has double spaces or odd line breaks in the raw data)."""
    return " ".join((text or "").split())


def _format_recipe_instructions(text: str) -> str:
    """TheMealDB's strInstructions field comes back as multiple steps separated
    by raw newlines, and each step can itself contain commas (e.g. 'Using a
    fork, poke holes...'). Left as-is, those embedded newlines/commas make the
    cell look "shredded" into extra columns/rows the moment it's opened in
    Excel or any tool that isn't a strict CSV parser. This rebuilds the whole
    thing as ONE clean, single-line string — "Step 1: ... | Step 2: ..." —
    with no embedded newlines at all, so it can never break formatting no
    matter what opens it."""
    if not text:
        return ""
    # Split on any run of newlines (MealDB uses \r\n, often doubled between steps).
    raw_steps = [s.strip() for s in re.split(r"[\r\n]+", text) if s.strip()]
    # Strip any "STEP 1", "Step 1:" etc. the source text already included, so
    # we can renumber consistently instead of ending up with "Step 1: Step 1 ...".
    cleaned_steps = []
    for step in raw_steps:
        step = re.sub(r"^(step\s*\d+[:.)]?\s*)", "", step, flags=re.IGNORECASE)
        step = _clean_ingredient_line(step)  # collapse internal whitespace too
        if step:
            cleaned_steps.append(step)
    return " | ".join(f"Step {i}: {s}" for i, s in enumerate(cleaned_steps, 1))


def _format_recipe_ingredients(ingredients: list) -> str:
    """Same idea as instructions above: one clean line, semicolon-separated,
    with whitespace normalized on every individual ingredient."""
    cleaned = [_clean_ingredient_line(ing) for ing in (ingredients or [])]
    return "; ".join(i for i in cleaned if i)


def _build_fusion_row(result: dict, usda_api_key: str, openweather_api_key: str,
                       ai_api_key: str = "", ai_model_name: str = "", ai_provider: str = "") -> dict:
    """Look up every relevant extra API for one result and return a flat dict of
    fused fields (excluding image/caption/etc, which the caller already has).
    Shared by both the CSV and XLSX exporters so the two never drift apart.

    ai_api_key/ai_model_name/ai_provider are optional — pass them to enable an
    AI-estimated nutrition fallback (clearly labelled "(AI estimate)") for
    Food/Vegetable/Fruit results where USDA genuinely has no match, so every
    food row still gets real calorie/protein/carb/fat numbers instead of
    being left blank."""
    name = result.get("NAME", "")
    category_lower = result.get("CATEGORY", "").lower()
    row = {}

    if not name:
        return row

    wiki = wikipedia_lookup(name, "en")
    if wiki:
        row["wikipedia_summary"] = wiki.get("extract", "")
        row["wikipedia_url"] = wiki.get("url", "")

    if any(k in category_lower for k in ("food", "vegetable", "fruit")):
        nutrition = usda_nutrition_lookup(name, usda_api_key)
        if not nutrition and ai_api_key:
            nutrition = ai_estimate_nutrition(name, ai_api_key, ai_model_name, ai_provider)
        if nutrition:
            n = nutrition.get("nutrients", {})
            row["nutrition_calories"] = n.get("Calories", "")
            row["nutrition_protein"] = n.get("Protein", "")
            row["nutrition_carbs"] = n.get("Carbs", "")
            row["nutrition_fat"] = n.get("Fat", "")

    if "food" in category_lower:
        recipe = mealdb_recipe_lookup(name)
        if recipe:
            row["recipe_ingredients"] = _format_recipe_ingredients(recipe.get("ingredients", []))
            row["recipe_instructions"] = _format_recipe_instructions(recipe.get("instructions", ""))

    if any(k in category_lower for k in ("animal", "flower")):
        species = gbif_species_lookup(name)
        if species:
            row["gbif_scientific_name"] = species.get("scientificName", "")
            row["gbif_kingdom"] = species.get("kingdom", "")
            row["gbif_family"] = species.get("family", "")
            row["gbif_status"] = species.get("status", "")

    if any(k in category_lower for k in ("place", "landmark", "building")) and openweather_api_key:
        weather, _error = openweather_lookup(name, openweather_api_key)
        if weather:
            row["weather_description"] = weather.get("description", "")
            row["weather_temp_c"] = weather.get("temp_c", "")

    return row


def export_history_to_csv(history: list, usda_api_key: str = "DEMO_KEY", openweather_api_key: str = "",
                           progress_callback=None, ai_api_key: str = "", ai_model_name: str = "",
                           ai_provider: str = "") -> bytes:
    """Save each history entry's image to IMAGES_DIR and return one combined CSV
    (as bytes) with one row per entry, referencing each image by filename —
    plain CSV can't embed pictures, so use the XLSX export for that."""
    buf = io.StringIO()
    # Excel decides how to split CSV columns using your Windows "list separator"
    # regional setting — in many locales that's NOT a comma, so double-clicking a
    # plain comma-CSV can dump everything into one column. This "sep=," line is a
    # Microsoft-specific directive that forces Excel to use commas regardless of
    # regional settings, whenever the file is opened directly (not imported).
    buf.write("sep=,\n")
    writer = csv.DictWriter(buf, fieldnames=CSV_COLUMNS)
    writer.writeheader()

    total = len(history)
    for i, entry in enumerate(history):
        if progress_callback:
            progress_callback(i, total)

        result = entry["result"]
        filename = _save_history_image(entry["image_bytes"])

        row = {
            "image_filename": filename,
            "caption": result.get("CAPTION", ""),
            "category": result.get("CATEGORY", ""),
            "name": result.get("NAME", ""),
            "confidence": result.get("CONFIDENCE", ""),
            "description": result.get("DESCRIPTION", ""),
            "details": result.get("DETAILS", ""),
            "fun_fact": result.get("FUN_FACT", ""),
        }
        row.update(_build_fusion_row(result, usda_api_key, openweather_api_key, ai_api_key, ai_model_name, ai_provider))
        writer.writerow(row)

    return buf.getvalue().encode("utf-8-sig")  # utf-8-sig so Excel opens accents/emoji correctly


# Column order matching the fusion Excel export exactly, as requested for ML use.
ML_CSV_COLUMNS = [
    "Image", "Caption", "Category", "Name", "Confidence",
    "Description", "Details", "Fun Fact",
    "Wikipedia Summary", "Wikipedia URL",
    "Calories", "Protein", "Carbs", "Fat",
    "Scientific Name", "Kingdom", "Family", "Conservation Status",
    "Recipe Ingredients", "Recipe Instructions",
    "Weather", "Temp (C)",
]


def _extract_number(value, default: float = 0.0) -> float:
    """Pull the leading numeric part out of a value like '52.0 KCAL' or '87%'
    and return it as a real float — ML algorithms need actual numbers, not
    text with units glued on. Returns `default` (0.0 unless overridden) when
    nothing numeric is found, so every cell in a numeric column is always a
    real number and never an empty/NaN cell."""
    if value is not None:
        text = str(value).strip()
        if text:
            num_chars = []
            for ch in text:
                if ch.isdigit() or ch in ".-":
                    num_chars.append(ch)
                else:
                    break
            num_str = "".join(num_chars).strip(".")
            if num_str and num_str not in ("-", "."):
                try:
                    return float(num_str)
                except ValueError:
                    pass
    return default


_FAKE_NULL_VALUES = {
    "not applicable", "n/a", "na", "none", "null", "unknown", "not available", "-",
}


def _clean_ml_text(value, default: str = "Unknown") -> str:
    """Make a text field safe and clean for an ML CSV: turn placeholder
    "missing" strings like 'Not Applicable' into `default` ("Unknown" unless
    overridden) so every cell is consistently filled rather than empty, and
    collapse any embedded newlines/tabs into single spaces so every row
    stays one clean line in the raw file."""
    if value is None:
        return default
    text = str(value).strip()
    if not text or text.lower().rstrip(".!") in _FAKE_NULL_VALUES:
        return default
    # Collapse embedded newlines/tabs/carriage returns and repeated whitespace.
    text = " ".join(text.split())
    return text


def export_history_to_ml_csv(history: list, usda_api_key: str = "DEMO_KEY", openweather_api_key: str = "",
                              progress_callback=None, ai_api_key: str = "", ai_model_name: str = "",
                              ai_provider: str = "") -> bytes:
    """Build a CSV specifically structured for feeding into a machine-learning
    pipeline (e.g. pandas -> scikit-learn): numeric columns (Confidence,
    Calories, Protein, Carbs, Fat, Temp) contain real numbers, not text with
    units attached; every text field is a single clean line with no stray
    'Not Applicable'-style placeholders; and every row is one complete,
    ML-ready training example. The Image column holds a filename reference
    (saved into IMAGES_DIR) since a CSV cell can't hold actual pixel data —
    load images separately in your ML code using that filename if you need
    the pixels themselves."""
    buf = io.StringIO()
    buf.write("sep=,\n")
    writer = csv.DictWriter(buf, fieldnames=ML_CSV_COLUMNS)
    writer.writeheader()

    total = len(history)
    for i, entry in enumerate(history):
        if progress_callback:
            progress_callback(i, total)

        result = entry["result"]
        filename = _save_history_image(entry["image_bytes"])
        fusion = _build_fusion_row(result, usda_api_key, openweather_api_key, ai_api_key, ai_model_name, ai_provider)
        category_lower = result.get("CATEGORY", "").lower()

        # "Not Applicable" = this field genuinely doesn't apply to this photo's
        # category (e.g. Weather for a Food photo) — expected, not a problem.
        # "Unknown" = the field DOES apply, but the matching external API
        # (GBIF / TheMealDB / OpenWeatherMap / USDA) didn't return a match or
        # the lookup failed. Collapsing both into one label made it impossible
        # to tell "nothing to look up" apart from "the lookup failed" — so
        # they're kept separate here.
        is_food = "food" in category_lower
        is_nutrition_cat = any(k in category_lower for k in ("food", "vegetable", "fruit"))
        is_taxonomy_cat = any(k in category_lower for k in ("animal", "flower"))
        is_place_cat = any(k in category_lower for k in ("place", "landmark", "building"))

        def _cat_text(value, applicable: bool) -> str:
            return _clean_ml_text(value, default=("Unknown" if applicable else "Not Applicable"))

        writer.writerow({
            "Image": filename,
            "Caption": _clean_ml_text(result.get("CAPTION", "")),
            "Category": _clean_ml_text(result.get("CATEGORY", "")),
            "Name": _clean_ml_text(result.get("NAME", "")),
            "Confidence": _extract_number(result.get("CONFIDENCE", "")),
            "Description": _clean_ml_text(result.get("DESCRIPTION", "")),
            "Details": _clean_ml_text(result.get("DETAILS", "")),
            "Fun Fact": _clean_ml_text(result.get("FUN_FACT", "")),
            "Wikipedia Summary": _clean_ml_text(fusion.get("wikipedia_summary", ""), default="Unknown"),
            "Wikipedia URL": _clean_ml_text(fusion.get("wikipedia_url", ""), default="Unknown"),
            "Calories": _extract_number(fusion.get("nutrition_calories", "")),
            "Protein": _extract_number(fusion.get("nutrition_protein", "")),
            "Carbs": _extract_number(fusion.get("nutrition_carbs", "")),
            "Fat": _extract_number(fusion.get("nutrition_fat", "")),
            "Scientific Name": _cat_text(fusion.get("gbif_scientific_name", ""), is_taxonomy_cat),
            "Kingdom": _cat_text(fusion.get("gbif_kingdom", ""), is_taxonomy_cat),
            "Family": _cat_text(fusion.get("gbif_family", ""), is_taxonomy_cat),
            "Conservation Status": _cat_text(fusion.get("gbif_status", ""), is_taxonomy_cat),
            "Recipe Ingredients": _cat_text(fusion.get("recipe_ingredients", ""), is_food),
            "Recipe Instructions": _cat_text(fusion.get("recipe_instructions", ""), is_food),
            "Weather": _cat_text(fusion.get("weather_description", ""), is_place_cat),
            "Temp (C)": _extract_number(fusion.get("weather_temp_c", "")),
        })

    return buf.getvalue().encode("utf-8-sig")


def export_history_to_xlsx(history: list, usda_api_key: str = "DEMO_KEY", openweather_api_key: str = "",
                            progress_callback=None, ai_api_key: str = "", ai_model_name: str = "",
                            ai_provider: str = "") -> bytes:
    """Build a real Excel workbook with the actual photo embedded in each row,
    plus every fused field in its own column — this is what makes it 'look like
    an actual dataset' when opened in Excel, since plain CSV can't hold images."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SnapKnow Data"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")

    for col_idx, header in enumerate(XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # Column widths: image column narrower/fixed, text columns wide enough to be readable.
    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, len(XLSX_HEADERS) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 28
    ws.freeze_panes = "A2"

    # Row height is in POINTS, image width/height are in PIXELS — a 120px image is
    # ~90pt tall at 96 DPI, so give the row a bit more than that to avoid the
    # image visually overlapping the row below it.
    THUMB_PX = 120
    ROW_HEIGHT = 105

    total = len(history)
    for i, entry in enumerate(history):
        if progress_callback:
            progress_callback(i, total)

        result = entry["result"]
        image_bytes = entry["image_bytes"]
        excel_row = i + 2  # row 1 is the header
        ws.row_dimensions[excel_row].height = ROW_HEIGHT

        # Embed a resized thumbnail (not the full-size photo) so the file stays a
        # reasonable size and every row reads at a consistent, tidy height.
        try:
            thumb_bytes = resize_image_bytes(image_bytes, max_dimension=THUMB_PX)
            img_stream = io.BytesIO(thumb_bytes)
            xl_img = XLImage(img_stream)
            xl_img.width = THUMB_PX
            xl_img.height = THUMB_PX
            ws.add_image(xl_img, f"A{excel_row}")
        except Exception:
            logger.error("Could not embed thumbnail into XLSX export", exc_info=True)
            ws.cell(row=excel_row, column=1, value="(image failed to embed)")

        fusion = _build_fusion_row(result, usda_api_key, openweather_api_key, ai_api_key, ai_model_name, ai_provider)
        values = [
            None,  # column A holds the embedded image, not text
            result.get("CAPTION", ""),
            result.get("CATEGORY", ""),
            result.get("NAME", ""),
            result.get("CONFIDENCE", ""),
            result.get("DESCRIPTION", ""),
            result.get("DETAILS", ""),
            result.get("FUN_FACT", ""),
            fusion.get("wikipedia_summary", ""),
            fusion.get("wikipedia_url", ""),
            fusion.get("nutrition_calories", ""),
            fusion.get("nutrition_protein", ""),
            fusion.get("nutrition_carbs", ""),
            fusion.get("nutrition_fat", ""),
            fusion.get("gbif_scientific_name", ""),
            fusion.get("gbif_kingdom", ""),
            fusion.get("gbif_family", ""),
            fusion.get("gbif_status", ""),
            fusion.get("recipe_ingredients", ""),
            fusion.get("recipe_instructions", ""),
            fusion.get("weather_description", ""),
            fusion.get("weather_temp_c", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            if col_idx == 1:
                continue  # leave the image column's cell value empty
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = wrap_top

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def category_style(category_text: str) -> dict:
    text = (category_text or "").lower()
    for key, style in CATEGORY_STYLES.items():
        if key in text:
            return style
    return DEFAULT_STYLE


# ---------------------------------------------------------------------------
# API #1 — Wikipedia REST API (no key needed)
# Verified encyclopedia summary + source link for whatever was identified.
# ---------------------------------------------------------------------------
def wikipedia_lookup(name: str, wiki_lang: str = "en"):
    if not name:
        return None
    title = name.split(",")[0].split("(")[0].strip()
    if not title:
        return None
    url = f"https://{wiki_lang}.wikipedia.org/api/rest_v1/page/summary/{urllib.parse.quote(title)}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "SnapKnow/1.0 (student project)"})
        with urllib.request.urlopen(req, timeout=6) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if data.get("type") == "disambiguation":
            return None
        extract = data.get("extract")
        page_url = data.get("content_urls", {}).get("desktop", {}).get("page")
        if not extract:
            return None
        return {"extract": extract, "url": page_url}
    except Exception:
        logger.info(f"Wikipedia lookup found nothing for '{name}' ({wiki_lang})", exc_info=True)
        return None


# ---------------------------------------------------------------------------
# API #2 — USDA FoodData Central API (free key, or DEMO_KEY for light testing)
# Real nutrition facts for Food / Vegetable / Fruit results.
# ---------------------------------------------------------------------------
_SEARCH_STOPWORDS = {
    "with", "and", "of", "in", "on", "at", "for", "a", "an", "the", "to",
    "topped", "served", "style", "recipe", "homemade", "fresh", "some",
}


def _narrowing_candidates(base_query: str) -> list:
    """Build a list of progressively shorter/more-generic search queries from a
    full AI-generated name, for APIs (USDA, TheMealDB) whose databases are
    small/exact-match and rarely contain a full composed name verbatim.
    e.g. "Belgian Waffles with Syrup" -> tries the full name, then
    "Belgian Waffles" and "Waffles" (filler words like "with" filtered out
    first so a real, searchable word is what actually gets tried)."""
    words = base_query.split()
    meaningful = [w for w in words if w.lower().strip(".,!") not in _SEARCH_STOPWORDS]
    if not meaningful:
        meaningful = words

    candidates = [base_query]
    if len(meaningful) > 1:
        candidates.append(" ".join(meaningful[-2:]))
        candidates.append(meaningful[-1])
        candidates.append(meaningful[0])
    elif meaningful:
        candidates.append(meaningful[0])

    seen = set()
    return [c.strip(".,!") for c in candidates
            if c.strip(".,!") and not (c.strip(".,!").lower() in seen or seen.add(c.strip(".,!").lower()))]


def _usda_search_once(query: str, api_key: str):
    """One raw call to USDA FoodData Central's search endpoint. Returns the
    first food dict or None — split out so usda_nutrition_lookup can retry
    with shorter/more generic queries."""
    url = (
        "https://api.nal.usda.gov/fdc/v1/foods/search?"
        + urllib.parse.urlencode({"query": query, "pageSize": 3, "api_key": api_key})
    )
    with urllib.request.urlopen(url, timeout=6) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    foods = data.get("foods") or []
    return foods[0] if foods else None


def usda_nutrition_lookup(name: str, api_key: str = "DEMO_KEY"):
    if not name:
        return None
    base_query = name.split(",")[0].split("(")[0].strip()
    if not base_query:
        return None

    # USDA FoodData Central is mostly raw ingredients and generic foods, not
    # composed dish names — "Belgian Waffles with Syrup" won't match anything,
    # but "Waffles" often will. Retry with progressively shorter queries.
    candidates = _narrowing_candidates(base_query)

    food = None
    for candidate in candidates:
        try:
            food = _usda_search_once(candidate, api_key)
        except Exception:
            logger.info(f"USDA FoodData Central lookup failed for '{candidate}'", exc_info=True)
            food = None
        if food:
            break

    if not food:
        logger.info(f"USDA FoodData Central found nothing usable for '{name}' after trying {candidates}")
        return None

    nutrients = {}
    for n in food.get("foodNutrients", []):
        label = n.get("nutrientName", "")
        value = n.get("value")
        unit = n.get("unitName", "")
        if value is None:
            continue
        for key, match in [("Calories", "Energy"), ("Protein", "Protein"),
                            ("Carbs", "Carbohydrate"), ("Fat", "Total lipid")]:
            if match.lower() in label.lower() and key not in nutrients:
                nutrients[key] = f"{value} {unit}"
    if not nutrients:
        return None
    return {"description": food.get("description", base_query), "nutrients": nutrients}


def ai_estimate_nutrition(name: str, api_key: str, model_name: str, provider: str):
    """When USDA genuinely has no match (even after narrowing the query), ask
    the AI for a best-effort per-serving estimate instead of leaving the CSV
    blank. Always labelled "(AI estimate)" in the returned text so it's never
    confused with verified USDA data. Returns None only if the AI call itself
    fails or gives back nothing parseable."""
    if not api_key:
        return None
    prompt = (
        f"Estimate typical nutrition per serving for \"{name}\". Reply with "
        "EXACTLY four lines and nothing else, plain numbers only (no units, "
        "no ranges, no words):\nCALORIES: <number>\nPROTEIN: <number>\n"
        "CARBS: <number>\nFAT: <number>"
    )
    try:
        raw = text_completion(prompt, api_key, model_name, provider).strip()
    except Exception:
        logger.info(f"AI nutrition estimate failed for '{name}'", exc_info=True)
        return None

    values = {}
    for line in raw.splitlines():
        if ":" not in line:
            continue
        label, _, rest = line.partition(":")
        label = label.strip().upper()
        if label in ("CALORIES", "PROTEIN", "CARBS", "FAT"):
            num = _extract_number(rest, default=None)
            if num is not None:
                values[label] = num
    if not values:
        return None
    return {
        "description": f"{name} (AI estimate)",
        "nutrients": {
            "Calories": f"{values.get('CALORIES', 0)} kcal (AI estimate)",
            "Protein": f"{values.get('PROTEIN', 0)} g (AI estimate)",
            "Carbs": f"{values.get('CARBS', 0)} g (AI estimate)",
            "Fat": f"{values.get('FAT', 0)} g (AI estimate)",
        },
    }


# ---------------------------------------------------------------------------
# API #3 — GBIF Species API (no key needed)
# Real taxonomy + conservation status for Animal / Flower results.
# ---------------------------------------------------------------------------
def gbif_species_lookup(name: str):
    if not name:
        return None
    query = name.split(",")[0].split("(")[0].strip()
    if not query:
        return None

    # Step 1 — species/match: GBIF's fast, strict fuzzy-match against scientific
    # names. Works great for "Lion", "Sunflower", etc. but fails for anything
    # more specific than species level (e.g. "Golden Retriever" is a dog BREED,
    # not a separate GBIF taxon — only "Canis lupus familiaris" exists there).
    url = "https://api.gbif.org/v1/species/match?" + urllib.parse.urlencode({"name": query})
    try:
        with urllib.request.urlopen(url, timeout=6) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if data.get("matchType") != "NONE" and data.get("scientificName"):
            return {
                "scientificName": data.get("scientificName"),
                "kingdom": data.get("kingdom"),
                "phylum": data.get("phylum"),
                "family": data.get("family"),
                "rank": data.get("rank"),
                "status": data.get("status"),
            }
    except Exception:
        logger.info(f"GBIF species/match found no result for '{name}'", exc_info=True)

    # Step 2 — species/search: a broader full-text search that also checks
    # vernacular/common names, not just scientific ones. This is what actually
    # catches breed/variety-level names by falling back to the parent species
    # GBIF DOES track (e.g. "Golden Retriever" search often surfaces
    # "Canis lupus" / "Canis familiaris" via its vernacular name index).
    try:
        search_url = "https://api.gbif.org/v1/species/search?" + urllib.parse.urlencode(
            {"q": query, "limit": 5}
        )
        with urllib.request.urlopen(search_url, timeout=6) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        for result in data.get("results", []):
            if result.get("scientificName") and result.get("rank") in (
                "SPECIES", "GENUS", "FAMILY", "SUBSPECIES", "VARIETY"
            ):
                return {
                    "scientificName": result.get("scientificName"),
                    "kingdom": result.get("kingdom"),
                    "phylum": result.get("phylum"),
                    "family": result.get("family"),
                    "rank": result.get("rank"),
                    "status": result.get("taxonomicStatus") or result.get("status"),
                }
    except Exception:
        logger.info(f"GBIF species/search also found no result for '{name}'", exc_info=True)

    return None


# ---------------------------------------------------------------------------
# API #4 — Unsplash API (free key from unsplash.com/developers)
# A couple of reference photos of the identified subject, for comparison.
# ---------------------------------------------------------------------------
def unsplash_reference_images(name: str, access_key: str, count: int = 3):
    if not name or not access_key:
        return []
    query = name.split(",")[0].split("(")[0].strip()
    if not query:
        return []
    url = (
        "https://api.unsplash.com/search/photos?"
        + urllib.parse.urlencode({"query": query, "per_page": count})
    )
    try:
        req = urllib.request.Request(url, headers={"Authorization": f"Client-ID {access_key}"})
        with urllib.request.urlopen(req, timeout=6) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        results = data.get("results") or []
        return [
            {"url": r["urls"]["small"], "credit": r["user"]["name"], "link": r["links"]["html"]}
            for r in results if r.get("urls")
        ]
    except Exception:
        logger.warning(f"Unsplash reference image lookup failed for '{name}'", exc_info=True)
        return []


# ---------------------------------------------------------------------------
# API #5 — QR Code Generator API (no key needed, goqr.me)
# A scannable QR code linking to the Wikipedia article, embedded in the PDF.
# ---------------------------------------------------------------------------
def generate_qr_code_bytes(data: str, size: int = 200):
    if not data:
        return None
    url = "https://api.qrserver.com/v1/create-qr-code/?" + urllib.parse.urlencode(
        {"size": f"{size}x{size}", "data": data}
    )
    try:
        with urllib.request.urlopen(url, timeout=6) as resp:
            return resp.read()
    except Exception:
        logger.warning(f"QR code generation failed for data='{data[:60]}...'", exc_info=True)
        return None


# ---------------------------------------------------------------------------
# API #6 — TheMealDB API (no key needed)
# A real recipe (ingredients + steps) for Food results.
# ---------------------------------------------------------------------------
def _mealdb_search_once(query: str):
    """One raw call to TheMealDB's search endpoint. Returns the first meal dict
    or None — split out so mealdb_recipe_lookup can retry with shorter queries."""
    if not query:
        return None
    url = "https://www.themealdb.com/api/json/v1/1/search.php?" + urllib.parse.urlencode({"s": query})
    with urllib.request.urlopen(url, timeout=6) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    meals = data.get("meals")
    return meals[0] if meals else None


def mealdb_recipe_lookup(dish_name: str):
    if not dish_name:
        return None
    base_query = dish_name.split(",")[0].split("(")[0].strip()
    if not base_query:
        return None

    # TheMealDB only has a few hundred recipes, so an AI-generated name like
    # "Homemade Vegetable Stir Fry" rarely matches anything verbatim. Retry
    # with progressively shorter/more generic queries built from the
    # meaningful (non-filler) words in the name.
    candidates = _narrowing_candidates(base_query)

    meal = None
    matched_query = base_query
    for candidate in candidates:
        try:
            meal = _mealdb_search_once(candidate)
        except Exception:
            logger.info(f"TheMealDB lookup failed for '{candidate}'", exc_info=True)
            meal = None
        if meal:
            matched_query = candidate
            break

    if not meal:
        logger.info(f"TheMealDB found no recipe for '{dish_name}' after trying {candidates}")
        return None

    if matched_query != base_query:
        logger.info(f"TheMealDB matched '{dish_name}' via narrowed query '{matched_query}'")

    ingredients = []
    for i in range(1, 21):
        ing = meal.get(f"strIngredient{i}")
        measure = meal.get(f"strMeasure{i}")
        if ing and ing.strip():
            line = ing.strip()
            if measure and measure.strip():
                line = f"{measure.strip()} {line}"
            ingredients.append(line)
    return {
        "name": meal.get("strMeal", base_query),
        "category": meal.get("strCategory", ""),
        "area": meal.get("strArea", ""),
        "ingredients": ingredients,
        "instructions": meal.get("strInstructions", ""),
        "thumbnail": meal.get("strMealThumb"),
    }


# ---------------------------------------------------------------------------
# API #7 — OpenWeatherMap API (free key from openweathermap.org)
# Current weather for Place results.
# ---------------------------------------------------------------------------
# A place identified as e.g. "Eiffel Tower" or "Taj Mahal" isn't a city, so
# OpenWeatherMap's direct weather-by-name endpoint (which only knows cities)
# rejects it with a 404. This maps well-known landmarks to the city they're
# actually in, so weather can still be looked up for them.
LANDMARK_TO_CITY = {
    "eiffel tower": "Paris", "taj mahal": "Agra", "statue of liberty": "New York",
    "great wall of china": "Beijing", "colosseum": "Rome", "colosseum rome": "Rome",
    "sydney opera house": "Sydney", "golden gate bridge": "San Francisco",
    "machu picchu": "Cusco", "christ the redeemer": "Rio de Janeiro",
    "burj khalifa": "Dubai", "mount fuji": "Tokyo", "niagara falls": "Niagara Falls",
    "grand canyon": "Grand Canyon Village", "stonehenge": "Salisbury",
    "big ben": "London", "leaning tower of pisa": "Pisa", "acropolis": "Athens",
    "buckingham palace": "London", "louvre": "Paris", "louvre museum": "Paris",
    "empire state building": "New York", "space needle": "Seattle",
    "sagrada familia": "Barcelona", "petra": "Wadi Musa", "angkor wat": "Siem Reap",
    "mount rushmore": "Keystone",
}


def _openweather_request(query: str, api_key: str):
    """One raw call to OpenWeatherMap's current-weather-by-name endpoint.
    Raises on failure — split out so openweather_lookup can retry with a
    resolved city name if the first query (a landmark name) is rejected."""
    url = "https://api.openweathermap.org/data/2.5/weather?" + urllib.parse.urlencode(
        {"q": query, "appid": api_key, "units": "metric"}
    )
    with urllib.request.urlopen(url, timeout=6) as resp:
        return json.loads(resp.read().decode("utf-8"))


def openweather_lookup(place_name: str, api_key: str):
    """Returns (weather_dict, error_message). Exactly one will be None, so the
    caller can show a useful reason instead of a generic 'nothing found'."""
    if not place_name:
        return None, "No place name to look up."
    if not api_key:
        return None, "No OpenWeatherMap key set."
    query = place_name.split(",")[0].split("(")[0].strip()
    if not query:
        return None, "No place name to look up."

    def _try(q):
        try:
            data = _openweather_request(q, api_key)
        except urllib.error.HTTPError as e:
            return None, e.code
        except Exception as e:
            return None, str(e)
        if str(data.get("cod")) != "200":
            return None, data.get("message", "Unknown error from OpenWeatherMap.")
        return data, None

    data, err = _try(query)

    # If the direct query failed because it isn't a recognised city (404),
    # and it matches a known landmark, retry using that landmark's actual city.
    if data is None and err == 404:
        resolved_city = LANDMARK_TO_CITY.get(query.strip().lower())
        if resolved_city:
            logger.info(f"OpenWeatherMap: '{query}' isn't a city — retrying as '{resolved_city}'")
            data, err2 = _try(resolved_city)
            if data is not None:
                err = None

    if data is None:
        if err == 401:
            logger.warning(f"OpenWeatherMap rejected the API key (401) for query '{query}'")
            return None, ("OpenWeatherMap rejected the key (401 Unauthorized). New keys can take "
                           "up to ~2 hours after signup to activate — try again shortly.")
        if err == 404:
            logger.info(f"OpenWeatherMap found no city matching '{query}'")
            return None, f"No city found matching \"{query}\". Weather lookup needs a city/place name, not a landmark or building name."
        if isinstance(err, int):
            logger.warning(f"OpenWeatherMap returned HTTP {err} for query '{query}'")
            return None, f"OpenWeatherMap error (HTTP {err})."
        logger.warning(f"Could not reach OpenWeatherMap for query '{query}': {err}")
        return None, f"Could not reach OpenWeatherMap: {err}"

    weather = (data.get("weather") or [{}])[0]
    main = data.get("main", {})
    return {
        "location": data.get("name", query),
        "description": weather.get("description", "").title(),
        "temp_c": main.get("temp"),
        "feels_like_c": main.get("feels_like"),
        "humidity": main.get("humidity"),
    }, None


def ai_fallback(prompt: str, api_key: str, model_name: str, provider: str):
    """When a real external API has no data for this item, ask the AI to fill
    that specific gap instead of just showing 'not found'. Always returns a
    string, clearly meant to be labeled as an estimate rather than verified data."""
    if not api_key:
        return None
    try:
        return text_completion(prompt, api_key, model_name, provider).strip()
    except Exception:
        logger.error("AI fallback completion failed", exc_info=True)
        return None


HERO_SVG = """
<svg width="120" height="120" viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg">
  <rect x="10" y="30" width="100" height="72" rx="10" fill="#1F2937"/>
  <rect x="42" y="18" width="36" height="20" rx="4" fill="#1F2937"/>
  <circle cx="75" cy="66" r="20" fill="#F76707"/>
  <circle cx="75" cy="66" r="12" fill="#FFF3E0"/>
  <circle cx="100" cy="46" r="4" fill="#F2A93B"/>
</svg>
"""


def inject_theme():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

        :root {
            --bg: #F8FAFC;
            --surface: #FFFFFF;
            --surface-alt: #F1F5F9;
            --ink: #0F172A;
            --ink-dim: #64748B;
            --primary: #2563EB;
            --primary-dark: #1D4ED8;
            --accent: #0EA5E9;
            --success: #16A34A;
            --warning: #D97706;
            --border: #E2E8F0;
        }

        html, body, [data-testid="stAppViewContainer"], .stApp {
            background-color: var(--bg) !important;
            color: var(--ink) !important;
            font-family: 'Inter', sans-serif;
        }

        [data-testid="stHeader"] { background-color: transparent; }

        [data-testid="stSidebar"] {
            background-color: var(--surface) !important;
            border-right: 1px solid var(--border);
        }
        [data-testid="stSidebar"] * { color: var(--ink) !important; }
        [data-testid="stSidebar"] input, [data-testid="stSidebar"] select,
        [data-testid="stSidebar"] [data-baseweb="select"] > div {
            background-color: var(--surface-alt) !important;
            border-radius: 8px !important;
            border: 1px solid var(--border) !important;
        }
        [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
            gap: 0.5rem !important;
        }
        [data-testid="stSidebar"] hr {
            margin: 0.4rem 0 !important;
            border-top: 1px solid var(--border) !important;
        }

        h1, h2, h3 { font-family: 'Inter', sans-serif !important; color: var(--ink) !important; font-weight: 700 !important; }

        [data-testid="stAppViewContainer"] .main .block-container { padding-top: 2rem; max-width: 1200px; }

        /* ---------- Hero: clean bordered banner, subtle accent bar ---------- */
        .hero-postcard {
            background-color: var(--surface);
            border: 1px solid var(--border);
            border-left: 4px solid var(--primary);
            border-radius: 10px;
            padding: 1.8rem 2.1rem;
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 1.6rem;
            box-shadow: 0 1px 3px rgba(15, 23, 42, 0.06);
        }
        .hero-tape { display: none; }
        .hero-text { max-width: 62ch; }
        .hero-eyebrow {
            font-family: 'Inter', sans-serif;
            font-weight: 600;
            font-size: 0.78rem;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            color: var(--primary);
            margin-bottom: 0.4rem;
        }
        .hero-title {
            font-family: 'Inter', sans-serif;
            font-weight: 800;
            font-size: 2rem;
            letter-spacing: -0.02em;
            color: var(--ink);
            margin: 0 0 0.4rem 0;
        }
        .hero-sub { color: var(--ink-dim); font-size: 0.95rem; }

        [data-testid="stButton"] button, [data-testid="stDownloadButton"] button {
            background: var(--primary) !important;
            color: #ffffff !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: 600 !important;
            font-family: 'Inter', sans-serif !important;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.08);
            transition: background 0.12s ease, box-shadow 0.12s ease;
        }
        [data-testid="stButton"] button:hover, [data-testid="stDownloadButton"] button:hover {
            background: var(--primary-dark) !important;
            box-shadow: 0 2px 6px rgba(37, 99, 235, 0.25);
        }
        [data-testid="stButton"] button:active, [data-testid="stDownloadButton"] button:active {
            background: var(--primary-dark) !important;
        }

        [data-testid="stTabs"] [role="tab"] {
            font-family: 'Inter', sans-serif;
            font-weight: 600;
            color: var(--ink-dim);
        }
        [data-testid="stTabs"] [aria-selected="true"] {
            color: var(--primary) !important;
            border-bottom: 2px solid var(--primary) !important;
        }

        [data-testid="stExpander"] {
            background-color: var(--surface) !important;
            border: 1px solid var(--border) !important;
            border-radius: 10px !important;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
        }

        [data-testid="stFileUploaderDropzone"], [data-testid="stAudioInput"] {
            background-color: var(--surface-alt) !important;
            border: 1.5px dashed var(--border) !important;
            border-radius: 10px !important;
        }
        [data-testid="stFileUploaderDropzone"] *, [data-testid="stAudioInput"] *,
        [data-testid="stFileUploaderDropzoneInstructions"],
        [data-testid="stFileUploaderDropzoneInstructions"] span,
        [data-testid="stFileUploaderDropzoneInstructions"] small {
            color: var(--ink) !important;
            font-family: 'Inter', sans-serif !important;
            opacity: 1 !important;
        }
        [data-testid="stFileUploaderDropzoneInstructions"] small { color: var(--ink-dim) !important; }
        [data-testid="stFileUploaderDropzone"] svg { fill: var(--primary) !important; opacity: 0.9; }
        [data-testid="stBaseButton-secondary"], [data-testid="stFileUploaderDropzone"] button {
            background: var(--surface) !important;
            color: var(--primary) !important;
            border: 1px solid var(--primary) !important;
            border-radius: 8px !important;
            font-family: 'Inter', sans-serif !important;
            font-weight: 600 !important;
        }

        [data-testid="stAudioInput"] {
            background-color: var(--surface-alt) !important;
        }
        [data-testid="stAudioInput"] div {
            background-color: transparent !important;
        }
        [data-testid="stAudioInput"] span, [data-testid="stAudioInput"] p {
            color: var(--ink) !important;
            opacity: 1 !important;
            font-family: 'Inter', sans-serif !important;
        }
        [data-testid="stAudioInput"] svg, [data-testid="stAudioInput"] path,
        [data-testid="stAudioInput"] circle {
            fill: var(--ink) !important;
            opacity: 0.7 !important;
        }
        [data-testid="stAudioInput"] canvas {
            filter: invert(0.75) contrast(1.3);
        }
        [data-testid="stAudioInput"] button {
            background: var(--primary) !important;
            border: none !important;
            border-radius: 50% !important;
        }
        [data-testid="stAudioInput"] button svg, [data-testid="stAudioInput"] button path {
            fill: #ffffff !important;
            opacity: 1 !important;
        }

        [data-testid="stSidebarCollapsedControl"], [data-testid="collapsedControl"] {
            background-color: var(--primary) !important;
            border: none !important;
            border-radius: 50% !important;
            box-shadow: 0 2px 6px rgba(37, 99, 235, 0.35);
        }
        [data-testid="stSidebarCollapsedControl"] svg, [data-testid="collapsedControl"] svg {
            fill: #ffffff !important;
        }

        [data-testid="stTextInput"] input {
            background-color: var(--surface) !important;
            color: var(--ink) !important;
            border: 1px solid var(--border) !important;
            border-radius: 8px !important;
        }

        hr { border-top: 1px solid var(--border) !important; }

        [data-testid="stCaptionContainer"], .stCaption { color: var(--ink-dim) !important; }

        [data-testid="stAlert"] {
            border-radius: 8px !important;
            background-color: #FFFBEB !important;
            border: 1px solid #FDE68A !important;
        }
        [data-testid="stAlert"] * { color: var(--ink) !important; opacity: 1 !important; }

        /* ---------- Result card: clean report-style panel ---------- */
        .polaroid-card {
            background-color: var(--surface);
            border: 1px solid var(--border);
            border-radius: 10px;
            padding: 1.3rem 1.5rem 1.6rem 1.5rem;
            margin-bottom: 1.2rem;
            box-shadow: 0 1px 3px rgba(15, 23, 42, 0.06);
        }
        .polaroid-sticker {
            display: inline-block;
            font-family: 'Inter', sans-serif; font-weight: 600; font-size: 0.7rem;
            letter-spacing: 0.04em; text-transform: uppercase;
            color: #ffffff; padding: 0.28rem 0.75rem; border-radius: 6px;
            margin-bottom: 0.6rem;
        }
        .polaroid-caption {
            font-family: 'Inter', sans-serif; font-weight: 700; font-size: 1.3rem;
            color: var(--ink); margin: 0.3rem 0 0 0; line-height: 1.35;
        }
        .polaroid-name {
            color: var(--ink-dim); font-size: 0.85rem; margin-bottom: 0.4rem;
        }
        .polaroid-section-label {
            font-family: 'Inter', sans-serif; font-size: 0.72rem; font-weight: 700;
            color: var(--primary); text-transform: uppercase; letter-spacing: 0.06em;
            margin-top: 0.9rem; margin-bottom: 0.25rem;
        }
        .polaroid-body { color: var(--ink); line-height: 1.6; font-size: 0.95rem; }
        .polaroid-funfact {
            background-color: #F0F9FF;
            border: 1px solid #BAE6FD;
            border-radius: 8px; padding: 0.65rem 0.9rem; margin-top: 1rem;
            color: #0369A1; font-family: 'Inter', sans-serif; font-size: 0.92rem;
        }
        .polaroid-answer {
            background-color: #F0FDF4; border: 1px solid #BBF7D0;
            padding: 0.7rem 0.9rem; border-radius: 8px; margin-top: 1rem; color: var(--ink);
        }
        .confidence-wrap {
            display: flex; align-items: center; gap: 0.5rem;
            margin: 0.6rem 0 0.2rem 0;
        }
        .confidence-track {
            flex: 1; height: 8px; background: var(--surface-alt);
            border-radius: 999px; overflow: hidden;
        }
        .confidence-fill { height: 100%; }
        .confidence-label {
            font-family: 'Inter', sans-serif; font-weight: 600; font-size: 0.78rem;
            white-space: nowrap;
        }

        /* ---------- History gallery row ---------- */
        .history-badge {
            display: inline-block; font-size: 0.68rem; font-weight: 600;
            text-transform: uppercase; letter-spacing: 0.03em; color: #ffffff;
            padding: 0.18rem 0.55rem; border-radius: 6px;
        }

        /* ---------- Collection gallery ---------- */
        .gallery-stats {
            font-family: 'Inter', sans-serif; font-size: 0.85rem; color: var(--ink-dim);
            margin-bottom: 0.8rem;
        }
        .gallery-card {
            background-color: var(--surface);
            border: 1px solid var(--border);
            border-radius: 8px;
            padding: 0.5rem 0.5rem 0.9rem 0.5rem;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
            text-align: center;
            margin-bottom: 0.4rem;
        }
        .gallery-card.rot-a, .gallery-card.rot-b, .gallery-card.rot-c { transform: none; }
        .gallery-card img {
            width: 100%; height: 90px; object-fit: cover; border-radius: 4px;
            margin-bottom: 0.4rem;
        }
        .gallery-caption {
            font-family: 'Inter', sans-serif; font-weight: 600; font-size: 0.88rem;
            color: var(--ink); line-height: 1.2;
        }

        /* ---------- "Ask Something" panel ---------- */
        .note-header {
            font-family: 'Inter', sans-serif;
            font-weight: 700;
            font-size: 1.05rem;
            color: var(--ink);
            margin: 0.6rem 0 0.5rem 0.1rem;
            display: inline-block;
        }
        .note-tape { display: none; }
        [data-testid="stVerticalBlockBorderWrapper"] {
            border: 1px solid var(--border) !important;
            border-radius: 10px !important;
            background-color: var(--surface) !important;
            box-shadow: 0 1px 3px rgba(15, 23, 42, 0.06) !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


inject_theme()

# ---------------------------------------------------------------------------
# Sidebar: provider + API key + model choice
# ---------------------------------------------------------------------------
st.sidebar.markdown("### ⚙️ Settings")

provider = st.sidebar.radio(
    "AI Provider",
    ["Google Gemini (free tier)", "Anthropic Claude (paid)", "Ollama (local, free)"],
    index=0,
    help="Gemini has a genuinely free tier (rate-limited, no card needed). Claude requires paid "
         "credits. Ollama runs entirely on your own machine — free, private, no internet needed "
         "once models are pulled, but needs Ollama installed and running locally.",
)

if provider.startswith("Google"):
    api_key = st.sidebar.text_input(
        "Gemini API Key",
        value=os.getenv("GEMINI_API_KEY", ""),
        type="password",
        help="Get a free key at https://aistudio.google.com/apikey — sign in with a Google account, "
             "click 'Create API key'. No credit card required for the free tier. Pre-filled "
             "automatically if GEMINI_API_KEY is set in a .env file.",
    )
    model_name = st.sidebar.selectbox(
        "Model",
        ["gemini-3.6-flash", "gemini-3.5-flash-lite", "gemini-2.5-flash", "gemini-3.1-pro-preview"],
        index=0,
        help="Flash and Flash-Lite are free-tier friendly (rate-limited). Pro Preview is paid-only "
             "but most capable — pick it only if you already have billing set up.",
    )
elif provider.startswith("Ollama"):
    api_key = st.sidebar.text_input(
        "Ollama Base URL",
        value=os.getenv("OLLAMA_BASE_URL", "http://localhost:11434"),
        help="The address Ollama is running on. Default is correct for a normal local install.",
    )
    model_name = st.sidebar.text_input(
        "Ollama Model",
        value=os.getenv("OLLAMA_VISION_MODEL", "llava"),
        help="Must be a vision-capable model you've already pulled, e.g. run "
             "`ollama pull llava` or `ollama pull llama3.2-vision` in a terminal first.",
    )
    st.sidebar.caption(
        "No API key needed — Ollama runs locally. Make sure the Ollama app/service is running "
        "and you've pulled a vision model before analyzing photos."
    )
else:
    api_key = st.sidebar.text_input(
        "Anthropic API Key",
        value=os.getenv("ANTHROPIC_API_KEY", ""),
        type="password",
        help="Get your key from https://console.anthropic.com/ . It is only used for this session "
             "and never stored. Pre-filled automatically if ANTHROPIC_API_KEY is set in a .env file.",
    )
    model_name = st.sidebar.selectbox(
        "Model",
        ["claude-sonnet-4-6", "claude-opus-4-6", "claude-haiku-4-6"],
        index=0,
        help="Sonnet is a good balance of speed/cost/quality. Opus is most capable, Haiku is fastest/cheapest.",
    )

st.sidebar.markdown("---")
language_choice = st.sidebar.selectbox(
    "🌐 Output Language",
    list(LANGUAGE_OPTIONS.keys()),
    index=0,
    help="Captions, descriptions, details, fun facts, and read-aloud speech will all use this "
         "language. Changing this after analyzing a photo automatically re-explains it in the "
         "new language.",
)
tts_lang = LANGUAGE_OPTIONS[language_choice]

st.sidebar.markdown("---")
with st.sidebar.expander("🔌 Extra APIs (optional)"):
    st.caption(
        "Wikipedia, GBIF, TheMealDB, and QR code generation work with no key. USDA, Unsplash, "
        "and OpenWeatherMap need a free key to unlock nutrition facts, reference photos, and "
        "weather — leave blank to skip those. All three are auto-filled from a .env file if set."
    )
    usda_api_key = st.text_input(
        "USDA FoodData Central API key",
        value=os.getenv("USDA_API_KEY", ""),
        type="password",
        help="Free key: https://fdc.nal.usda.gov/api-key-signup — or leave blank to use the "
             "shared DEMO_KEY (very low rate limit, fine for quick testing).",
    )
    unsplash_access_key = st.text_input(
        "Unsplash Access Key",
        value=os.getenv("UNSPLASH_ACCESS_KEY", ""),
        type="password",
        help="Free key: https://unsplash.com/developers — leave blank to skip reference photos.",
    )
    openweather_api_key = st.text_input(
        "OpenWeatherMap API key",
        value=os.getenv("OPENWEATHER_API_KEY", ""),
        type="password",
        help="Free key: https://openweathermap.org/api — leave blank to skip weather for Place results.",
    )
if not usda_api_key:
    usda_api_key = "DEMO_KEY"

st.sidebar.markdown("---")
with st.sidebar.expander("🧠 RAG Settings (Ollama)"):
    st.caption(
        "Powers the 'Ask Your Collection' feature — a real retrieval-augmented pipeline "
        "(LangChain + Chroma vector DB + Ollama) that answers questions across everything "
        "you've analyzed. Runs fully locally; needs Ollama installed and running."
    )
    rag_base_url = st.text_input(
        "Ollama Base URL",
        value=os.getenv("OLLAMA_BASE_URL", "http://localhost:11434"),
        key="rag_base_url",
    )
    rag_embed_model = st.text_input(
        "Embedding model",
        value=os.getenv("OLLAMA_EMBED_MODEL", "nomic-embed-text"),
        help="Pull it first: `ollama pull nomic-embed-text`",
    )
    rag_chat_model = st.text_input(
        "Answer generation model",
        value=os.getenv("OLLAMA_RAG_MODEL", "llama3.2"),
        help="Any Ollama chat model works, e.g. `ollama pull llama3.2`",
    )

st.sidebar.markdown("---")
st.sidebar.caption(
    "Image captioning & knowledge run on your chosen provider above. Text-to-speech uses gTTS "
    "and speech-to-text uses Google's free web speech service — both work regardless of which "
    "provider you pick. History is saved locally next to app.py so it survives a refresh."
)

hero_html = f"""
<div class="hero-postcard">
    <div class="hero-tape left"></div>
    <div class="hero-tape right"></div>
    <div class="hero-text">
        <div class="hero-eyebrow">Snap it. Know it.</div>
        <div class="hero-title">📸 SnapKnow</div>
        <div class="hero-sub">
            Upload any photo — an animal, a flower, a dish, anything — and get it identified,
            explained, read aloud, and saved as a PDF report.
        </div>
    </div>
    <div>{HERO_SVG}</div>
</div>
"""
st.markdown(
    "\n".join(line.strip() for line in hero_html.strip().split("\n")),
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------
if "history" not in st.session_state:
    st.session_state.history = load_history_from_disk()
if "last_result" not in st.session_state:
    st.session_state.last_result = None
if "last_image_bytes" not in st.session_state:
    st.session_state.last_image_bytes = None
if "qa_history" not in st.session_state:
    st.session_state.qa_history = []
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0
if "last_language" not in st.session_state:
    st.session_state.last_language = language_choice
if "translation_cache" not in st.session_state:
    # Caches {(image_hash, language): parsed_result} so switching back to a
    # language you've already viewed for this image is instant, no API call.
    st.session_state.translation_cache = {}


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------
def image_hash(image_bytes: bytes) -> str:
    """Short fingerprint used as a cache key so re-analyzing the same image in a
    language you've already viewed doesn't need a fresh API call."""
    return hashlib.md5(image_bytes).hexdigest()


def text_to_speech(text: str, lang: str = "en") -> str:
    """Generate an mp3 file from text and return its path."""
    tts = gTTS(text=text, lang=lang)
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".mp3")
    tts.save(tmp_file.name)
    return tmp_file.name


def speech_to_text(audio_bytes: bytes):
    """Transcribe recorded audio bytes (wav) to text using free Google Web Speech API."""
    recognizer = sr.Recognizer()
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".wav")
    tmp_file.write(audio_bytes)
    tmp_file.close()
    try:
        with sr.AudioFile(tmp_file.name) as source:
            audio_data = recognizer.record(source)
            text = recognizer.recognize_google(audio_data)
        return text
    except sr.UnknownValueError:
        logger.info("Speech recognition could not understand the audio")
        return None
    except Exception as e:
        logger.warning("Speech-to-text request failed", exc_info=True)
        return f"ERROR: {e}"
    finally:
        try:
            os.unlink(tmp_file.name)
        except OSError:
            logger.debug(f"Could not delete temp audio file {tmp_file.name}", exc_info=True)


def generate_pdf(image_bytes: bytes, parsed: dict, filename: str = "image_report.pdf") -> str:
    """Build a nicely formatted PDF report with the image + all extracted info."""
    tmp_img = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
    tmp_img.close()  # release the handle immediately — Windows can't delete an open file later otherwise
    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    img.save(tmp_img.name, "JPEG")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "AI Image Analysis Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(
        0, 8, datetime.now().strftime("Generated on %Y-%m-%d %H:%M"),
        new_x="LMARGIN", new_y="NEXT", align="C",
    )
    pdf.ln(4)

    pdf.image(tmp_img.name, x=55, w=100)
    # pdf.image() can leave the cursor at the image's right edge; explicitly
    # reset it to the left margin below the image before writing more text,
    # otherwise multi_cell() has no horizontal room left and raises
    # "Not enough horizontal space to render a single character". Compute the
    # exact rendered height from the image's own aspect ratio (width is fixed
    # at 100mm above) so the cursor lands right below the image, not on it.
    img_w_px, img_h_px = img.size
    rendered_height_mm = 100 * (img_h_px / img_w_px)
    new_y = pdf.get_y() + rendered_height_mm + 4
    if new_y > pdf.page_break_trigger:
        pdf.add_page()
        new_y = pdf.get_y()
    pdf.set_xy(pdf.l_margin, new_y)

    def safe(text: str) -> str:
        # fpdf2's built-in Helvetica font only supports Latin-1; replace any
        # characters outside that range (smart quotes, emoji, etc.) so the
        # PDF doesn't crash or render garbled boxes.
        return text.encode("latin-1", "replace").decode("latin-1")

    pdf.set_font("Helvetica", "B", 14)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(0, 10, safe(parsed.get("CAPTION", "")))

    for field, label in [
        ("CATEGORY", "Category"),
        ("NAME", "Name"),
        ("DESCRIPTION", "Description"),
        ("DETAILS", "Details"),
        ("FUN_FACT", "Fun Fact"),
    ]:
        if parsed.get(field):
            pdf.set_x(pdf.l_margin)
            pdf.set_font("Helvetica", "B", 12)
            pdf.multi_cell(0, 8, label + ":")
            pdf.set_x(pdf.l_margin)
            pdf.set_font("Helvetica", "", 11)
            pdf.multi_cell(0, 7, safe(parsed[field]))
            pdf.ln(2)

    # QR code linking to a Wikipedia search for the identified subject — built
    # directly from the name (no extra Wikipedia API call here) so PDF
    # generation, which currently re-runs on every page render, stays fast.
    qr_name = parsed.get("NAME", "").strip()
    if qr_name:
        qr_url = "https://en.wikipedia.org/wiki/Special:Search?search=" + urllib.parse.quote(qr_name)
        qr_bytes = generate_qr_code_bytes(qr_url, size=150)
        if qr_bytes:
            tmp_qr = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            tmp_qr.close()
            with open(tmp_qr.name, "wb") as f:
                f.write(qr_bytes)
            if pdf.get_y() + 35 > pdf.page_break_trigger:
                pdf.add_page()
            pdf.ln(4)
            pdf.set_x(pdf.l_margin)
            pdf.set_font("Helvetica", "I", 9)
            pdf.multi_cell(0, 6, safe("Scan to look up more:"))
            pdf.image(tmp_qr.name, x=pdf.l_margin, w=25)
            try:
                os.unlink(tmp_qr.name)
            except OSError:
                logger.debug(f"Could not delete temp QR file {tmp_qr.name}", exc_info=True)

    out_path = os.path.join(tempfile.gettempdir(), filename)
    pdf.output(out_path)
    try:
        os.unlink(tmp_img.name)
    except OSError:
        logger.debug(f"Could not delete temp image file {tmp_img.name}", exc_info=True)
    return out_path


def _load_font(size: int, bold: bool = False):
    """Try a few common TrueType fonts, fall back to PIL's basic bitmap font if
    none are found — the card still works either way, just less polished."""
    from PIL import ImageFont
    candidates = (
        ["arialbd.ttf", "Arial Bold.ttf", "DejaVuSans-Bold.ttf"] if bold
        else ["arial.ttf", "Arial.ttf", "DejaVuSans.ttf"]
    )
    for name in candidates:
        try:
            return ImageFont.truetype(name, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _wrap_text(draw, text, font, max_width):
    """Word-wrap text to fit within max_width, used since PIL doesn't wrap text itself."""
    words = text.split()
    lines, current = [], ""
    for word in words:
        trial = (current + " " + word).strip()
        if draw.textlength(trial, font=font) <= max_width:
            current = trial
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def generate_share_card(image_bytes: bytes, parsed: dict) -> bytes:
    """Build a polished, branded square image (photo + caption + fun fact) that
    the user can download and post on Instagram/WhatsApp/etc."""
    from PIL import ImageDraw

    W, H = 1080, 1080
    photo_h = 680
    style = category_style(parsed.get("CATEGORY", ""))
    accent = style["text"]

    card = Image.new("RGB", (W, H), "#FFFFFF")
    draw = ImageDraw.Draw(card)

    # Photo, cropped to fill the top section
    photo = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    src_ratio = photo.width / photo.height
    dst_ratio = W / photo_h
    if src_ratio > dst_ratio:
        new_h = photo.height
        new_w = int(new_h * dst_ratio)
        x0 = (photo.width - new_w) // 2
        photo = photo.crop((x0, 0, x0 + new_w, new_h))
    else:
        new_w = photo.width
        new_h = int(new_w / dst_ratio)
        y0 = (photo.height - new_h) // 2
        photo = photo.crop((0, y0, new_w, y0 + new_h))
    photo = photo.resize((W, photo_h), Image.LANCZOS)
    card.paste(photo, (0, 0))

    # Category sticker
    tag_font = _load_font(30, bold=True)
    tag_text = f"{style['emoji']} {parsed.get('CATEGORY', 'Unclassified').upper()}"
    tag_w = draw.textlength(tag_text, font=tag_font) + 50
    draw.rounded_rectangle([40, photo_h - 60, 40 + tag_w, photo_h - 10], radius=20, fill=accent)
    draw.text((65, photo_h - 52), tag_text, font=tag_font, fill="#FFFFFF")

    # Bottom text section
    caption_font = _load_font(52, bold=True)
    body_font = _load_font(30)
    brand_font = _load_font(26, bold=True)

    y = photo_h + 40
    for line in _wrap_text(draw, parsed.get("CAPTION", ""), caption_font, W - 80)[:3]:
        draw.text((40, y), line, font=caption_font, fill="#1F2937")
        y += 62

    y += 10
    fun_fact = parsed.get("FUN_FACT", "")
    if fun_fact:
        draw.rounded_rectangle([40, y, W - 40, H - 90], radius=16, fill="#FFE49E")
        ty = y + 24
        for line in _wrap_text(draw, "✨ " + fun_fact, body_font, W - 130)[:5]:
            draw.text((65, ty), line, font=body_font, fill="#7A5200")
            ty += 40

    draw.text((40, H - 55), "📸 SnapKnow", font=brand_font, fill=accent)

    buf = io.BytesIO()
    card.save(buf, format="PNG")
    return buf.getvalue()


def generate_combined_pdf(history: list, filename: str = "snapknow_all_history.pdf") -> str:
    """Build one PDF containing every analyzed image + its info, one entry per page."""

    def safe(text: str) -> str:
        return text.encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    for entry in history:
        parsed = entry["result"]
        image_bytes = entry["image_bytes"]

        tmp_img = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
        tmp_img.close()
        img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
        img.save(tmp_img.name, "JPEG")

        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.multi_cell(0, 10, safe(parsed.get("CAPTION", "")))
        pdf.image(tmp_img.name, x=55, w=100)

        img_w_px, img_h_px = img.size
        rendered_height_mm = 100 * (img_h_px / img_w_px)
        new_y = pdf.get_y() + rendered_height_mm + 4
        if new_y > pdf.page_break_trigger:
            pdf.add_page()
            new_y = pdf.get_y()
        pdf.set_xy(pdf.l_margin, new_y)

        for field, label in [
            ("CATEGORY", "Category"),
            ("NAME", "Name"),
            ("DESCRIPTION", "Description"),
            ("DETAILS", "Details"),
            ("FUN_FACT", "Fun Fact"),
        ]:
            if parsed.get(field):
                pdf.set_x(pdf.l_margin)
                pdf.set_font("Helvetica", "B", 12)
                pdf.multi_cell(0, 8, label + ":")
                pdf.set_x(pdf.l_margin)
                pdf.set_font("Helvetica", "", 11)
                pdf.multi_cell(0, 7, safe(parsed[field]))
                pdf.ln(2)

        try:
            os.unlink(tmp_img.name)
        except OSError:
            logger.debug(f"Could not delete temp image file {tmp_img.name}", exc_info=True)

    out_path = os.path.join(tempfile.gettempdir(), filename)
    pdf.output(out_path)
    return out_path


# ---------------------------------------------------------------------------
# If the person changes the language dropdown after already having a result
# on screen, translate the existing text into the new language. Cache hits
# apply instantly here; anything needing a real API call is deferred to a
# flag checked inside col2, so the rest of the page (upload box, history)
# keeps rendering normally instead of the whole app going blank while we wait.
# ---------------------------------------------------------------------------
if "pending_translation" not in st.session_state:
    st.session_state.pending_translation = False

if (
    language_choice != st.session_state.last_language
    and st.session_state.last_image_bytes is not None
    and st.session_state.last_result is not None
):
    st.session_state.last_language = language_choice
    cache_key = (image_hash(st.session_state.last_image_bytes), language_choice)
    cached = st.session_state.translation_cache.get(cache_key)
    if cached:
        st.session_state.last_result = cached
        st.session_state.qa_history = []
        st.session_state.pending_translation = False
    elif api_key:
        st.session_state.pending_translation = True
    else:
        st.session_state.pending_translation = False
        st.warning("Enter your API key in the sidebar to switch languages on an existing result.")


# ---------------------------------------------------------------------------
# Main layout
# ---------------------------------------------------------------------------
col1, col2 = st.columns([1, 1.3])

with col1:
    uploaded_file = st.file_uploader(
        "Upload an image",
        type=["jpg", "jpeg", "png", "webp"],
        key=f"file_uploader_{st.session_state.uploader_key}",
    )
    if uploaded_file:
        image_bytes = resize_image_bytes(uploaded_file.getvalue())
        st.session_state.last_image_bytes = image_bytes
        st.image(image_bytes, caption="Uploaded image", use_container_width=True)

        if st.button("🔍 Analyze Image", type="primary", use_container_width=True):
            if not api_key:
                st.error("Please enter your API key in the sidebar.")
            else:
                with st.spinner(f"Analyzing image with {provider.split(' (')[0]}..."):
                    media_type = get_media_type(uploaded_file.name)
                    try:
                        raw = analyze_image(
                            image_bytes, media_type, api_key, model_name, provider, language_choice
                        )
                        parsed = parse_response(raw)
                        st.session_state.last_result = parsed
                        st.session_state.qa_history = []
                        st.session_state.translation_cache[
                            (image_hash(image_bytes), language_choice)
                        ] = parsed
                        st.session_state.history.append(
                            {"result": parsed, "image_bytes": image_bytes}
                        )
                        save_history_to_disk(st.session_state.history)
                    except Exception as e:
                        logger.error("Image analysis API call failed", exc_info=True)
                        st.error(f"Error calling API: {e}")

    elif st.session_state.last_image_bytes:
        # No fresh upload this run (e.g. the user picked an image from History),
        # but we still have an image loaded — show it so it isn't blank.
        st.image(st.session_state.last_image_bytes, caption="Selected image", use_container_width=True)
        st.caption("Loaded from history. Upload a new file above to analyze a different image.")

    st.divider()
    note_header_html = """
        <div class="note-header">
            <span class="note-tape"></span>
            💬 Ask Something
        </div>
        """
    st.markdown(
        "\n".join(line.strip() for line in note_header_html.strip().split("\n")),
        unsafe_allow_html=True,
    )

    with st.container(border=True):
        text_tab, voice_tab = st.tabs(["⌨️ Type", "🎤 Voice"])

        with text_tab:
            text_question = st.text_input(
                "Type your question about the image",
                placeholder="e.g. Is this safe for dogs to eat?",
                key="text_question_input",
            )
            if st.button("➡️ Ask", key="ask_text_button"):
                if not api_key:
                    st.error("Please enter your API key in the sidebar.")
                elif st.session_state.last_image_bytes is None:
                    st.error("Upload and analyze an image first.")
                elif not text_question.strip():
                    st.error("Type a question first.")
                else:
                    with st.spinner("Getting answer..."):
                        media_type = get_media_type("img.jpg")
                        raw = analyze_image(
                            st.session_state.last_image_bytes,
                            media_type,
                            api_key,
                            model_name,
                            provider,
                            language_choice,
                            extra_question=text_question.strip(),
                        )
                        parsed = parse_response(raw)
                        st.session_state.last_result = parsed
                        st.session_state.qa_history.append(
                            {"question": text_question.strip(), "answer": parsed.get("ANSWER", "")}
                        )

        with voice_tab:
            st.caption("Click the mic, allow browser microphone access, record your question, then click below.")
            audio_value = st.audio_input("Record your question")
            if audio_value and st.button("🗣️ Transcribe & Ask", key="ask_voice_button"):
                if not api_key:
                    st.error("Please enter your API key in the sidebar.")
                elif st.session_state.last_image_bytes is None:
                    st.error("Upload and analyze an image first.")
                else:
                    with st.spinner("Transcribing..."):
                        question = speech_to_text(audio_value.getvalue())
                    if not question:
                        st.error("Could not understand the audio, please try again.")
                    elif isinstance(question, str) and question.startswith("ERROR:"):
                        st.error(question)
                    else:
                        st.success(f"You asked: {question}")
                        with st.spinner("Getting answer..."):
                            media_type = get_media_type("img.jpg")
                            raw = analyze_image(
                                st.session_state.last_image_bytes,
                                media_type,
                                api_key,
                                model_name,
                                provider,
                                language_choice,
                                extra_question=question,
                            )
                            parsed = parse_response(raw)
                            st.session_state.last_result = parsed
                            st.session_state.qa_history.append(
                                {"question": question, "answer": parsed.get("ANSWER", "")}
                            )

with col2:
    if st.session_state.pending_translation:
        with st.spinner(f"Translating to {language_choice}..."):
            try:
                translated = translate_result_text(
                    st.session_state.last_result, language_choice, api_key, model_name, provider
                )
                cache_key = (image_hash(st.session_state.last_image_bytes), language_choice)
                st.session_state.last_result = translated
                st.session_state.translation_cache[cache_key] = translated
                st.session_state.qa_history = []
            except Exception as e:
                logger.error(f"Translation to {language_choice} failed", exc_info=True)
                st.error(f"Could not translate to {language_choice}: {e}")
            finally:
                st.session_state.pending_translation = False

    result = st.session_state.last_result
    if result:
        style = category_style(result.get("CATEGORY", ""))

        confidence_html = ""
        conf_raw = result.get("CONFIDENCE", "")
        try:
            conf_val = max(0, min(100, int("".join(ch for ch in conf_raw if ch.isdigit()))))
            conf_color = "#00D97E" if conf_val >= 75 else ("#FFC700" if conf_val >= 50 else "#FF3B3B")
            confidence_html = f"""
            <div class="confidence-wrap">
                <div class="confidence-track">
                    <div class="confidence-fill" style="width:{conf_val}%;background-color:{conf_color};"></div>
                </div>
                <span class="confidence-label" style="color:{conf_color};">{conf_val}% confident</span>
            </div>
            """
        except (ValueError, TypeError):
            logger.debug(f"Model returned no usable CONFIDENCE value: '{conf_raw}' — skipping the bar")

        card_html = f"""
        <div class="polaroid-card">
            <span class="polaroid-sticker" style="background-color:{style['bg']};color:{style['text']};">
                {style['emoji']} {result.get('CATEGORY', 'Unclassified')}
            </span>
            <div class="polaroid-caption">{result.get('CAPTION', '')}</div>
            <div class="polaroid-name">{result.get('NAME', '')}</div>
            {confidence_html}
            <div class="polaroid-section-label">Description</div>
            <div class="polaroid-body">{result.get('DESCRIPTION', '')}</div>
            <div class="polaroid-section-label">Details</div>
            <div class="polaroid-body">{result.get('DETAILS', '')}</div>
            <div class="polaroid-funfact">✨ {result.get('FUN_FACT', '')}</div>
            {f'<div class="polaroid-answer"><strong>💬 Answer:</strong> {result["ANSWER"]}</div>' if result.get('ANSWER') else ''}
        </div>
        """
        st.markdown(
            "\n".join(line.strip() for line in card_html.strip().split("\n")),
            unsafe_allow_html=True,
        )

        category_lower = result.get("CATEGORY", "").lower()
        name_for_lookup = result.get("NAME", "")

        if name_for_lookup:
            with st.expander(f"📖 Wikipedia summary for \"{name_for_lookup}\""):
                with st.spinner("Checking Wikipedia..."):
                    wiki = wikipedia_lookup(name_for_lookup, "en")
                if wiki:
                    st.write(wiki["extract"])
                    if wiki.get("url"):
                        st.markdown(f"[Read the full article →]({wiki['url']})")
                else:
                    st.caption("No matching Wikipedia article found — asking the AI instead:")
                    fallback = ai_fallback(
                        f"Give a short, factual, encyclopedia-style summary (2-3 sentences) of \"{name_for_lookup}\".",
                        api_key, model_name, provider,
                    )
                    if fallback:
                        st.warning(f"⚠️ AI-estimated (no verified Wikipedia article found):\n\n{fallback}")
                    else:
                        st.caption("Could not generate a fallback answer either — check your API key.")

        if name_for_lookup and any(k in category_lower for k in ("food", "vegetable", "fruit")):
            with st.expander(f"🥗 Nutrition facts for \"{name_for_lookup}\" (USDA)"):
                with st.spinner("Checking USDA FoodData Central..."):
                    nutrition = usda_nutrition_lookup(name_for_lookup, usda_api_key)
                if nutrition:
                    st.caption(f"Closest USDA match: {nutrition['description']}")
                    for k, v in nutrition["nutrients"].items():
                        st.markdown(f"**{k}:** {v}")
                else:
                    st.caption("No USDA nutrition match found — asking the AI instead:")
                    fallback = ai_fallback(
                        f"Estimate typical nutrition facts (calories, protein, carbs, fat, per "
                        f"100g) for \"{name_for_lookup}\". Keep it to a short list.",
                        api_key, model_name, provider,
                    )
                    if fallback:
                        st.warning(f"⚠️ AI-estimated (no verified USDA match found):\n\n{fallback}")
                    else:
                        st.caption("Could not generate a fallback answer either — check your API key.")

        if name_for_lookup and "food" in category_lower:
            with st.expander(f"🍳 Recipe for \"{name_for_lookup}\" (TheMealDB)"):
                with st.spinner("Looking up a recipe..."):
                    recipe = mealdb_recipe_lookup(name_for_lookup)
                if recipe:
                    if recipe.get("thumbnail"):
                        st.image(recipe["thumbnail"], width=200)
                    st.caption(f"{recipe['name']} — {recipe.get('area', '')} {recipe.get('category', '')}".strip())
                    st.markdown("**Ingredients:**")
                    for ing in recipe["ingredients"]:
                        st.markdown(f"- {ing}")
                    st.markdown("**Instructions:**")
                    st.write(recipe["instructions"])
                else:
                    st.caption("No matching recipe found on TheMealDB — asking the AI instead:")
                    fallback = ai_fallback(
                        f"Give a simple recipe for \"{name_for_lookup}\": a short ingredient list "
                        f"and numbered steps.",
                        api_key, model_name, provider,
                    )
                    if fallback:
                        st.warning(f"⚠️ AI-estimated (no verified TheMealDB match found):\n\n{fallback}")
                    else:
                        st.caption("Could not generate a fallback answer either — check your API key.")

        if name_for_lookup and any(k in category_lower for k in ("animal", "flower")):
            with st.expander(f"🔬 Species data for \"{name_for_lookup}\" (GBIF)"):
                with st.spinner("Checking GBIF species database..."):
                    species = gbif_species_lookup(name_for_lookup)
                if species:
                    st.markdown(f"**Scientific name:** {species.get('scientificName', '—')}")
                    st.markdown(f"**Kingdom:** {species.get('kingdom', '—')}  |  **Family:** {species.get('family', '—')}")
                    st.markdown(f"**Taxonomic rank:** {species.get('rank', '—')}  |  **Status:** {species.get('status', '—')}")
                else:
                    st.caption("No GBIF species match found — asking the AI instead:")
                    fallback = ai_fallback(
                        f"Give your best estimate of the scientific name, kingdom, family, and "
                        f"conservation status for \"{name_for_lookup}\". Keep it to a short list.",
                        api_key, model_name, provider,
                    )
                    if fallback:
                        st.warning(f"⚠️ AI-estimated (no verified GBIF match found):\n\n{fallback}")
                    else:
                        st.caption("Could not generate a fallback answer either — check your API key.")

        if name_for_lookup and any(k in category_lower for k in ("place", "landmark", "building")):
            with st.expander(f"🌤️ Current weather in \"{name_for_lookup}\" (OpenWeatherMap)"):
                weather, error = (None, "No OpenWeatherMap key set.")
                if openweather_api_key:
                    with st.spinner("Checking current weather..."):
                        weather, error = openweather_lookup(name_for_lookup, openweather_api_key)
                if weather:
                    st.markdown(f"**{weather['location']}** — {weather['description']}")
                    st.markdown(f"🌡️ {weather['temp_c']}°C (feels like {weather['feels_like_c']}°C)  |  💧 {weather['humidity']}% humidity")
                else:
                    st.caption(f"{error} Asking the AI for a general climate note instead:")
                    fallback = ai_fallback(
                        f"Live weather data isn't available for \"{name_for_lookup}\". Give a "
                        f"short, general note on the typical climate/season there instead, and "
                        f"make clear this is not live current weather.",
                        api_key, model_name, provider,
                    )
                    if fallback:
                        st.warning(f"⚠️ AI-estimated general climate note (not live weather):\n\n{fallback}")
                    else:
                        st.caption("Could not generate a fallback answer either — check your API key.")

        if name_for_lookup and unsplash_access_key:
            with st.expander(f"🖼️ Reference photos of \"{name_for_lookup}\" (Unsplash)"):
                with st.spinner("Fetching reference photos..."):
                    ref_images = unsplash_reference_images(name_for_lookup, unsplash_access_key)
                if ref_images:
                    ref_cols = st.columns(len(ref_images))
                    for c, img_info in zip(ref_cols, ref_images):
                        with c:
                            st.image(img_info["url"], use_container_width=True)
                            st.caption(f"Photo: {img_info['credit']}")
                else:
                    st.caption(
                        "No reference photos found on Unsplash for this term. (There's no honest "
                        "AI fallback here — the AI can't provide real photos of the actual subject, "
                        "only generate new images, which would misrepresent what's being shown.)"
                    )

        if st.session_state.qa_history:
            with st.expander(f"💬 Follow-up Q&A ({len(st.session_state.qa_history)})"):
                for qa in reversed(st.session_state.qa_history):
                    st.markdown(f"**Q:** {qa['question']}")
                    st.markdown(f"**A:** {qa['answer']}")
                    st.markdown("---")

        full_text = " ".join(
            [result.get(f, "") for f in ["CAPTION", "DESCRIPTION", "DETAILS", "FUN_FACT"]]
        )

        colA, colB, colC = st.columns(3)
        with colA:
            if st.button("🔊 Read Aloud"):
                with st.spinner("Generating speech..."):
                    audio_path = text_to_speech(full_text, lang=tts_lang)
                st.audio(audio_path, format="audio/mp3")
        with colB:
            if st.session_state.last_image_bytes:
                pdf_path = generate_pdf(st.session_state.last_image_bytes, result)
                with open(pdf_path, "rb") as f:
                    st.download_button(
                        "📄 Download PDF Report",
                        f,
                        file_name="image_report.pdf",
                        mime="application/pdf",
                    )
        with colC:
            if st.session_state.last_image_bytes:
                share_bytes = generate_share_card(st.session_state.last_image_bytes, result)
                st.download_button(
                    "📤 Share Card",
                    share_bytes,
                    file_name="snapknow_share_card.png",
                    mime="image/png",
                )
    else:
        st.info("Upload an image and click **Analyze Image** to see results here.")

st.divider()
with st.expander("🗂️ Collection Gallery", expanded=False):
    if not st.session_state.history:
        st.caption("No images analyzed yet.")
    else:
        hdr_col1, hdr_col2, hdr_col3, hdr_col4, hdr_col5 = st.columns(5)
        with hdr_col1:
            combined_pdf_path = generate_combined_pdf(st.session_state.history)
            with open(combined_pdf_path, "rb") as f:
                st.download_button(
                    "📚 Export All as PDF",
                    f,
                    file_name="snapknow_all_history.pdf",
                    mime="application/pdf",
                    key="export_all_history",
                )
        with hdr_col2:
            if st.button("🖼️ Prepare Fusion Excel", key="prepare_xlsx_history",
                         help="Real .xlsx file with the actual photo embedded in every row. Nutrition facts that USDA can't find are filled in with a clearly-labelled AI estimate instead of being left blank."):
                progress = st.progress(0.0, text="Fetching Wikipedia, nutrition, species, recipe, weather data...")

                def _update_progress_xlsx(i, total):
                    progress.progress((i + 1) / max(total, 1), text=f"Building row {i + 1} of {total}...")

                st.session_state.fusion_xlsx_bytes = export_history_to_xlsx(
                    st.session_state.history, usda_api_key, openweather_api_key, _update_progress_xlsx,
                    api_key, model_name, provider,
                )
                progress.empty()
            if st.session_state.get("fusion_xlsx_bytes"):
                st.download_button(
                    "⬇️ Download Fusion Excel",
                    st.session_state.fusion_xlsx_bytes,
                    file_name="snapknow_history.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_xlsx_history",
                )
        with hdr_col3:
            if st.button("📊 Prepare Fusion CSV", key="prepare_csv_history",
                         help="Plain-text data only, no embedded images. Nutrition facts that USDA can't find are filled in with a clearly-labelled AI estimate instead of being left blank."):
                progress = st.progress(0.0, text="Fetching Wikipedia, nutrition, species, recipe, weather data...")

                def _update_progress_csv(i, total):
                    progress.progress((i + 1) / max(total, 1), text=f"Fusing data for photo {i + 1} of {total}...")

                st.session_state.fusion_csv_bytes = export_history_to_csv(
                    st.session_state.history, usda_api_key, openweather_api_key, _update_progress_csv,
                    api_key, model_name, provider,
                )
                progress.empty()
            if st.session_state.get("fusion_csv_bytes"):
                st.download_button(
                    "⬇️ Download Fusion CSV",
                    st.session_state.fusion_csv_bytes,
                    file_name="snapknow_history.csv",
                    mime="text/csv",
                    key="download_csv_history",
                )
        with hdr_col4:
            if st.button("🤖 Prepare ML CSV", key="prepare_ml_csv_history",
                         help="Same columns as the Fusion Excel, but numeric fields (Confidence, Calories, Protein, Carbs, Fat, Temp) are clean numbers. Nutrition facts that USDA can't find are filled in with a clearly-labelled AI estimate instead of being left blank."):
                progress = st.progress(0.0, text="Fetching Wikipedia, nutrition, species, recipe, weather data...")

                def _update_progress_ml(i, total):
                    progress.progress((i + 1) / max(total, 1), text=f"Building row {i + 1} of {total}...")

                st.session_state.ml_csv_bytes = export_history_to_ml_csv(
                    st.session_state.history, usda_api_key, openweather_api_key, _update_progress_ml,
                    api_key, model_name, provider,
                )
                progress.empty()
            if st.session_state.get("ml_csv_bytes"):
                st.download_button(
                    "⬇️ Download ML CSV",
                    st.session_state.ml_csv_bytes,
                    file_name="snapknow_ml_dataset.csv",
                    mime="text/csv",
                    key="download_ml_csv_history",
                )
        with hdr_col5:
            if st.button("🗑️ Clear History", key="clear_history"):
                st.session_state.history = []
                save_history_to_disk(st.session_state.history)
                st.rerun()

        st.markdown("---")
        st.markdown("**🧠 Ask Your Collection** — retrieval-augmented Q&A across everything you've analyzed")
        rag_question = st.text_input(
            "Ask a question across your whole collection",
            placeholder="e.g. Which of my photos had the most calories? What animals have I identified?",
            key="rag_question_input",
            label_visibility="collapsed",
        )
        rag_col1, rag_col2 = st.columns([1, 1])
        with rag_col1:
            rag_ask_clicked = st.button("🔍 Ask", key="rag_ask_button")
        with rag_col2:
            rag_reindex_clicked = st.button("🔄 Rebuild Index", key="rag_reindex_button")

        if rag_reindex_clicked:
            with st.spinner("Embedding every photo in your collection into the vector store..."):
                try:
                    st.session_state.rag_vectordb = rag_engine.build_vectorstore(
                        st.session_state.history, rag_base_url, rag_embed_model
                    )
                    st.success(f"Index rebuilt from {len(st.session_state.history)} photos.")
                except ImportError:
                    st.error(
                        "RAG dependencies aren't installed. Run: "
                        "`pip install langchain langchain-ollama langchain-chroma langchain-core`"
                    )
                except Exception as e:
                    logger.error("Failed to build RAG vector store", exc_info=True)
                    st.error(
                        f"Could not reach Ollama at {rag_base_url}: {e}\n\n"
                        "Make sure Ollama is installed and running, and you've pulled the "
                        f"embedding model: `ollama pull {rag_embed_model}`"
                    )

        if rag_ask_clicked and rag_question.strip():
            if "rag_vectordb" not in st.session_state:
                st.warning("Click '🔄 Rebuild Index' first to embed your collection.")
            else:
                with st.spinner("Retrieving relevant photos and generating an answer..."):
                    try:
                        answer, sources = rag_engine.rag_query(
                            rag_question.strip(), st.session_state.rag_vectordb,
                            rag_base_url, rag_chat_model,
                        )
                        st.markdown(f"**Answer:** {answer}")
                        if sources:
                            with st.expander(f"📎 Based on {len(sources)} retrieved photo(s)"):
                                for d in sources:
                                    st.caption(f"• {d.metadata.get('caption', '(untitled)')}")
                    except ImportError:
                        st.error(
                            "RAG dependencies aren't installed. Run: "
                            "`pip install langchain langchain-ollama langchain-chroma langchain-core`"
                        )
                    except Exception as e:
                        logger.error("RAG query failed", exc_info=True)
                        st.error(f"Could not reach Ollama at {rag_base_url}: {e}")

        # Running collection stats, e.g. "🐾 3 Animals · 🌸 2 Flowers · 🍽️ 4 Dishes"
        counts = {}
        for h in st.session_state.history:
            cat = h["result"].get("CATEGORY", "Other").strip() or "Other"
            counts[cat] = counts.get(cat, 0) + 1
        stats_parts = []
        for cat, n in sorted(counts.items(), key=lambda kv: -kv[1]):
            emoji = category_style(cat)["emoji"]
            stats_parts.append(f"{emoji} {n} {cat}{'s' if n != 1 else ''}")
        st.markdown(f"<div class='gallery-stats'>{' &nbsp;·&nbsp; '.join(stats_parts)}</div>", unsafe_allow_html=True)

        rotations = ["rot-a", "rot-b", "rot-c"]
        items = list(reversed(st.session_state.history))
        cols_per_row = 4
        for row_start in range(0, len(items), cols_per_row):
            row_items = items[row_start:row_start + cols_per_row]
            grid_cols = st.columns(cols_per_row)
            for j, h in enumerate(row_items):
                i = row_start + j
                entry_result = h["result"]
                entry_image = h["image_bytes"]
                entry_style = category_style(entry_result.get("CATEGORY", ""))
                thumb_mime = guess_media_type_from_bytes(entry_image)
                thumb_b64 = base64.b64encode(entry_image).decode("ascii")
                rot = rotations[i % len(rotations)]
                with grid_cols[j]:
                    gallery_card_html = f"""
                        <div class="gallery-card {rot}">
                            <img src="data:{thumb_mime};base64,{thumb_b64}">
                            <div class="gallery-caption">{entry_style['emoji']} {entry_result.get('CAPTION', '')[:40]}</div>
                        </div>
                        """
                    st.markdown(
                        "\n".join(line.strip() for line in gallery_card_html.strip().split("\n")),
                        unsafe_allow_html=True,
                    )
                    if st.button("🔎 View", key=f"view_history_{i}", use_container_width=True):
                        st.session_state.last_result = entry_result
                        st.session_state.last_image_bytes = entry_image
                        st.session_state.qa_history = []
                        # Bump the uploader's key so the file_uploader widget resets to
                        # empty on rerun — otherwise it keeps re-displaying whatever was
                        # uploaded earlier instead of the historical image just selected.
                        st.session_state.uploader_key += 1
                        st.rerun()